# -*- coding: utf-8 -*-
'''
名称：Utils.py
功能：解析bugreport.log文件
'''
import re
import openpyxl
from openpyxl.cell.text import RichText
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import PatternFill, Alignment, Font
import main
from tools import properties_handler

# 通用的变量
words = properties_handler.Properties('D:\Documents\Desktop\开机时间工具\TimeCatch\keywords.properties').getProperties()
kernel_pattern = r"\d+\.\d{6}" # 匹配类似 "[    1.165322]", "[   13.371734]"的时间戳
time_pattern = r"\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3}"

def find_kw(text, kw_list):
    for kw in kw_list:
        if kw in text:
            return kw
    return None

'''
@params: texts 读取的log内容
return: kernel_dict, android_dict. 包含关键字和对应时间
'''
def get_keywords_and_timestamp(texts):
    android_keywords = [words['android'].get(i) for i in words['android']]
    kernel_keywords = [words['kernel'].get(i) for i in words['kernel']]
    kernel_keywords_mtk = [words['kernelmtk'].get(i) for i in words['kernelmtk']]
    android_dict = {key: value for key, value in zip(android_keywords, [-1] * len(android_keywords))}
    kernel_dict = {key: value for key, value in zip(kernel_keywords, [-1] * len(kernel_keywords))}
    kernel_mtk_dict = {key: value for key, value in zip(kernel_keywords_mtk, [-1] * len(kernel_keywords_mtk))}
    process_android = len(android_keywords)
    process_kernel = len(kernel_keywords)
    process_kernel_mtk = len(kernel_keywords_mtk)
    device_name = 'unknown'
    platform_name = 'unknown'
    for text in texts:
        if device_name == 'unknown' and words['device']['name'] in text:
            device_name = text.split(" ")[-1].strip('\n')
            continue

        if platform_name == 'unknown' and words['device']['platform'] in text:
            platform_name = text.split(" ")[-1].strip('\n').strip('"')
            if 'mt' in platform_name:
                platform_name = 'MTK'
                continue

        # kernel
        res = find_kw(text, kernel_keywords)
        timestamp = re.findall(kernel_pattern, text) # 非起始匹配
        if res is not None and timestamp is not None and process_kernel > 0:
            if kernel_dict[res] == -1:
                kernel_dict[res] = int(float(timestamp[0])*1000)
                kernel_keywords.remove(res)
                process_kernel -= 1
                continue

        # kernel_mtk
        res = find_kw(text, kernel_keywords_mtk)
        if res is not None and timestamp is not None and process_kernel_mtk > 0:
            if kernel_mtk_dict[res] == -1:
                kernel_mtk_dict[res] = int(float(timestamp[0])*1000)
                kernel_keywords_mtk.remove(res)
                process_kernel_mtk -= 1
                continue

        # android
        res = find_kw(text, android_keywords)
        timestamp = re.match(time_pattern, text)  # 起始匹配
        if res is not None and timestamp is not None and process_android > 0:
            if android_dict[res] == -1:
                android_dict[res] = int(text.split(' ')[-1].strip('\n'))
                android_keywords.remove(res)
                process_android -= 1
                continue
    # 将 boot_progress_start 挪到 kernel
    # kernel_dict[list(android_dict.keys())[0]] = android_dict[list(android_dict.keys())[0]]
    # android_dict.pop(list(android_dict.keys())[0])
    #print(kernel_mtk_dict)
    return kernel_dict, android_dict, [device_name, platform_name], kernel_mtk_dict

'''
@params: time_dict: 记录关键字和对应时间节点的字典对象
@return: duration_dict: 记录关键字和对应阶段耗时的字典
'''
def get_time_duration(time_dict):
    time_stamp_list = list(time_dict.values())
    # 异常值去除
    for idx, t in enumerate(time_stamp_list):
        if t == -1 and idx > 0:
            time_stamp_list[idx] = time_stamp_list[idx-1]
    duration_list = [a - b for a, b in zip(time_stamp_list, [0] + time_stamp_list[:-1])]
    duration_dict = {key: value for key, value in zip(time_dict.keys(), duration_list)}
    return duration_dict

'''
功能待定。。。
'''
def dict_to_xlsx(duration_dict_1, duration_dict_2, sheet_name, wb):
    file_path = words['path']['kernel_android']
    sheet_kernel_android = wb.active
    sheet_kernel_android.title = sheet_name

    #t = time.strftime("%m-%d_%H_%M", time.localtime())
    #wb.save(file_path + t + '.xlsx')


def dict_to_xlsx_contrast(sheet_obj, title, duration_dict_1, duration_dict_2):

    write_to_xlsx_column(title, sheet_obj, row_idx=1, column_start_idx='A')
    k1, d1 = duration_dict_1
    kernel_end_idx = len(k1)
    # k1_total = k1[list(k1.keys())[-1]] # kernel总耗时
    d1_total = d1[list(d1.keys())[-1]] # 开机总耗时
    k1.update(d1)
    d1_dur = get_time_duration(k1)
    write_to_xlsx_column(list(d1_dur.values()), sheet_obj, row_start_idx=2, column_idx='C')

    k2, d2 = duration_dict_2
    k2_total = k2[list(k2.keys())[-1]]
    d2_total = d2[list(d2.keys())[-1]]
    k2.update(d2)
    d2_dur = get_time_duration(k2)
    write_to_xlsx_column(list(d2_dur.values()), sheet_obj, row_start_idx=2, column_idx='D')

    keyword_list = list(k1.keys())
    write_to_xlsx_column(keyword_list, sheet_obj, row_start_idx=2, column_idx='B')

    # sheet_obj.insert_rows(2+kernel_end_idx)
    # write_to_xlsx_column(['kernel_total', k1_total, k2_total], sheet_obj, row_idx=2+kernel_end_idx, column_start_idx='B')
    write_to_xlsx_column(['总耗时', d1_total, d2_total], sheet_obj, row_idx=2+len(k1), column_start_idx='B')

    write_function_to_xlsx(sheet_obj, r_start_idx=2, r_end_idx=3+len(d2_dur),
                           c_start_idx=ord('C')-ord('A'), c_end_idx=ord('D')-ord('A'), signal='-')


    # 格式调整
    fill1 = PatternFill(fill_type='solid', start_color='ADD8E6', end_color='ADD8E6')  # 蔚蓝色
    fill2 = PatternFill(fill_type='solid', start_color='F0E68C', end_color='F0E68C')  # 黄色
    sheet_obj.cell(row=2, column=1).value = 'kernel'
    sheet_obj.cell(row=2, column=1).fill = fill1
    # sheet_obj.cell(row=2+kernel_end_idx, column=2).fill = fill2
    sheet_obj.merge_cells(range_string='A2:A'+str(2+kernel_end_idx-1))
    sheet_obj.merge_cells(range_string='A'+str(2+kernel_end_idx) + ':A'+str(1+len(d2_dur)))
    sheet_obj.cell(row=2+kernel_end_idx, column=1).value = 'android'
    sheet_obj.cell(row=2 + kernel_end_idx, column=1).fill = fill2
    # sheet_obj.cell(row=2+len(d2_dur) , column=2).fill = fill2

    # 设置 列 宽
    sheet_obj.column_dimensions['A'].width = 15
    sheet_obj.column_dimensions['B'].width = 37
    for column in ['C', 'D', 'E']:
        sheet_obj.column_dimensions[column].width = 12
    # 设置对齐方式
    cell_alignment(sheet_obj, Alignment(horizontal='center', vertical='center'), columns_start='A', columns_end='A')
    cell_alignment(sheet_obj, Alignment(horizontal='center'), columns_start='C', columns_end='E')


'''
# 设置单元格格式属性
@params: alignment : 格式对象
@return: None
'''
def cell_alignment(sheet_obj,  alignment, row_start=None, row_end=None, columns_start=None, columns_end=None):
    if row_start is None and columns_start is not None:
        i = 0
        while i+ord(columns_start) <= ord(columns_end):
            columns_name = chr(i+ord(columns_start))
            for j in range(1, 40):
                sheet_obj[columns_name+str(j)].alignment = alignment
            i += 1

'''
#   写入函数式
@param: signal : 符号（+，-，*，/）
'''
def write_function_to_xlsx(sheet_obj, r_start_idx, r_end_idx, c_start_idx, c_end_idx, signal, r_or_c=1):
    if r_or_c == 1:
        start_column_name = chr(c_start_idx + ord('A'))
        end_column_name = chr(c_end_idx + ord('A'))
        result_column_name = chr(ord(end_column_name)+1)
        if signal == '-':
            for i in range(r_end_idx-r_start_idx):
                cur_row = r_start_idx + i
                data1 = sheet_obj[start_column_name+str(cur_row)].value
                data2 = sheet_obj[end_column_name + str(cur_row)].value
                if type(data1)==int and type(data2)==int and data1-data2 < 300:
                    pass
                else:
                    sheet_obj.cell(row=cur_row, column=c_end_idx + 2).fill \
                        = PatternFill(fill_type='solid', start_color='FF7F50', end_color='FF4500')  # 橙红色
                sheet_obj[result_column_name+str(cur_row)]\
                    = '=' + start_column_name + str(cur_row) + signal + end_column_name + str(cur_row)

'''
# 写入内容到 sheet 对象中
@params: 
    data: 需要写入的数据，为列表对象。
    sheet_obj: 数据写入的 sheet 对象
'''
def write_to_xlsx_column(data, sheet_obj, row_start_idx=None, row_idx=None , column_start_idx=None,
                         column_idx=None, fill=None):
    # 按 行 写入， 给出 写入行序号row_idx，起始 列 名称 column_start_idx
    if row_idx is not None and column_start_idx is not None:
        for i in range(len(data)):
            sheet_obj.cell(row=row_idx, column=ord(column_start_idx) - ord('A')+1+i).value\
                = data[i]
        return
    # 按 列 写入，  给出 写入 列 名称 column_idx，起始 行 号 row_start_idx
    if row_start_idx is not None and column_idx is not None:
        for i in range(len(data)):
            sheet_obj.cell(row=row_start_idx+i, column=ord(column_idx) - ord('A')+1).value\
                = data[i] if data[i] != 0 or i == 0 else 'None'
        return

'''
# 柱状图绘制：根据时间节点绘制 测试机和对比机的耗时柱状图
@params: dict1, dict2 # {关键字：对应耗时}
'''
def general_Bar_chart(dict1, dict2, sheet_obj):
    bar_chart = BarChart()
    x1_axis = dict1.keys()
    y1_axis = dict1.values()
    x2_axis = dict2.keys()
    y2_axis = dict2.values()

    values = Reference(sheet_obj, min_col=3, min_row=1, max_col=4, max_row=len(dict1)+1)
    x_values = Reference(sheet_obj, min_col=2, min_row=2, max_col=2, max_row=len(dict2)+1)
    bar_chart.add_data(values, titles_from_data=True)
    #bar_chart.series(values, x_values, title="阶段耗时柱状图", titles_from_data=True)
    bar_chart.x_axis.title = '阶段'
    bar_chart.set_categories(x_values)
    bar_chart.y_axis.title = '耗时(ms)'
    bar_chart.width = len(dict1)*1
    bar_chart.height = 15
    bar_chart.gapWidth = 500
    # 柱状图中 每个柱子上方显示数值
    # data_labels = DataLabelList()
    # data_labels.showVal = True  # 显示数值
    # bar_chart.dataLabels = data_labels
    bar_chart.title = "阶段耗时柱状图"

    sheet_obj.add_chart(bar_chart, 'A'+str(len(x1_axis)+5))




'''
# excel 表格生成
@params1: file1, file2       #[kernel_dic, android_dict, (device_name, device_platform), kernel_dict]时间
                        节点字典（kernel，Android， kernel_mtk）, 设备代号以及平台
@params2: save_path          # excel表保存路径
'''
def xlsx_general(file_1, file_2, save_path):
    # d1 = load_data(file_1)
    # d2 = load_data(file_2)

    # d1_k, d1_b, d1_name = get_keywords_and_timestamp(d1)
    # d2_k, d2_b, d2_name = get_keywords_and_timestamp(d2)
    d1_k, d1_b, devcies_info1, kernel_mtk1 = file_1
    d2_k, d2_b, devcies_info2, kernel_mtk2 = file_2
    d1_name, d1_platform = devcies_info1
    d2_name, d2_platform = devcies_info2
    if d1_platform == d2_platform == 'MTK':
        d1_k.update(kernel_mtk1)
        d2_k.update(kernel_mtk2)
    wb = openpyxl.Workbook()
    sheet_kernel_android = wb.active
    sheet_kernel_android.title = 'kernel_android'
    dict_to_xlsx_contrast(sheet_kernel_android, ["阶段", "关键节点", d1_name, d2_name, "差值(ms)"],
                          [d1_k, d1_b], [d2_k, d2_b])
    general_Bar_chart(d1_k, d2_k, sheet_kernel_android) # 经过前面的过程，此时 d1_k已经包含了kernel和Android整个阶段的耗时
    wb.save(save_path)


def load_data(file_name):
    with open(file_name, 'Ur', encoding='utf-8') as f:
        return f.readlines()

if __name__ == '__main__':
    m = main.Main('D:/Documents/Desktop/AI/bugreport-warm.txt', 'D:/Documents/Desktop/AI/bugreport-V.txt')
    m.general_android_result()
    # write_function_to_xlsx(2, 8, 3, 4, '-', 1)
