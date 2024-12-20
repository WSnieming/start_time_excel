import re

import openpyxl
from openpyxl import Workbook
import pandas as pd
import numpy as np
import os
import sys
import logging
import datetime
from loguru import logger
import time

def load_log_file(filename):
    # //打开文件
    with open(filename, 'r', encoding='utf-8', errors='ignore') as f:
        lines = f.readlines()
        # 方法用于读取所有行(直到结束符
        # EOF)并返回列表，
        return lines

# //基于df 创建了表格；
def create_workshhet(df):
    # 设置了一下，说明
    df.insert(loc=0, column='Boot Step', value=None)
    # df['Time'] = df['Time'].astype(str)
    # df['Time'] = df['Time'].apply(lambda x: x.strip("[]"))
    # df['Time'] = df['Time'].apply(lambda x: x.strip("''"))

    df.to_excel('output.xlsx', index=False)

    wb = openpyxl.load_workbook('output.xlsx')
    ws = wb.active
    # 创建并合并kernel单元格
    # ws.insert_cols(1,amount = 1) row 行

    # 创建并合并Android单元格
    cell = ws.cell(row=2, column=1, value='Android')
    ws.merge_cells(start_row=2, start_column=1, end_row=21, end_column=1)

    cell = ws.cell(row=22, column=1, value='Kernel')
    ws.merge_cells(start_row=22, start_column=1, end_row=33, end_column=1)
    wb.save("output.xlsx")


from pandas.api.types import CategoricalDtype
def readkey(lines, name,file):
      # current_path =os.path.dirname(os.path.realpath(sys.executable))
      current_path = "./"
#     读取关键字
      coparefile = os.path.join(current_path, file)
      xls = pd.read_excel(coparefile);
      if name == "chuankou":
#           读取串口
          str=xls["串口log关键字"].tolist()
          df=find_keyword_1(lines,name,str);

          cat_size_order = CategoricalDtype(
                str,
                ordered=True
            )
          df['Keyword'] = df['Keyword'].astype(cat_size_order)
          newdf=df.sort_values('Keyword')

          # # 先设置为category：
          # df['Keyword'] = df['Keyword'].astype('category')
          # # 再设置category的顺序：
          # df['Keyword'].cat.set_categories(str)
      if name == "bugreport" or name=="bugcont":
          name="bugreport"
          str = xls["bugreport关键字"].tolist()
          df = find_keyword_1(lines, name, str);
          # 去除lsit  nan
          while np.nan in str:
               str.remove(np.nan)
          # 之后用new_list代替list_a
          cat_size_order = CategoricalDtype(
              str,
              ordered=True
          )
          df['Keyword'] = df['Keyword'].astype(cat_size_order)
          newdf = df.sort_values('Keyword')

      return  newdf;


def find_keyword_1(lines, name, str1):

    # //一共分为15步骤；
    time = []
    keyworldStr = []
    # 串口
    pattern1 = re.compile(r'(\d{2}:\d{2}:\d{2}.\d{3})')
    pattern2 = re.compile(r'(\d+\.?\d*)')

    pattern3 = re.compile(r'-?\d+\.?\d*e?-?\d*?')
    pattern4 = re.compile(r'(\d+)')
    keyworldStr=[]
    # lines=[	"行  8120: [Fri Nov 03 09:02:01.345 2023] [   10.898198][T1700275] init: [name:bootprof&]BOOTPROF:     10898.164102:INIT:Mount_END",
	# "行  9460: [Fri Nov 03 09:02:08.657 2023] [   18.217944][T1000275] init: [name:bootprof&]BOOTPROF:     18217.921273:INIT:Mount_END --late"]
    if name=="chuankou":
        for line in lines:
            # if any(substring in line for substring in str1 ):
            #     times = pattern1.findall(line)
            #     keyworldStr.append(substring)
            #     time.append(times)
            for str_name in str1[0:12]:
                if str_name in line:
                    times = pattern1.findall(line)[-1]
                    keyworldStr.append(str_name)
                    time.append(times)
                    break

            for str in str1[12:]:
                if (str=="INIT:Mount_START") & (str in line) & ("INIT:Mount_START --late" not in line):
                    times = pattern1.findall(line)[-1]
                    keyworldStr.append(str)
                    time.append(times)
                    break

                if line.find("INIT:Mount_START --late") != -1:
                    times = pattern1.findall(line)[-1]
                    keyworldStr.append("INIT:Mount_START --late")
                    time.append(times)
                    break

                if  (str=="INIT:Mount_END") & (str in line) & ("INIT:Mount_END --late" not in line):
                    times = pattern1.findall(line)[-1]
                    keyworldStr.append(str)
                    time.append(times)
                    break

                if  line.find("INIT:Mount_END --late") !=-1:
                    times = pattern1.findall(line)[-1]
                    keyworldStr.append("INIT:Mount_END --late")
                    time.append(times)
                    break

                if (str == "INIT:post-fs") & (str in line) & ("INIT:post-fs-data" not in line):
                    times = pattern1.findall(line)[-1]
                    keyworldStr.append(str)
                    time.append(times)
                    break

                if line.find("INIT:post-fs-data")!=-1:
                    times = pattern1.findall(line)[-1]
                    keyworldStr.append("INIT:post-fs-data")
                    time.append(times)
                    break;

                elif str in line:
                    times = pattern1.findall(line)[-1]
                    keyworldStr.append(str)
                    time.append(times)
                    break;
    if name=="bugreport":
        for line in lines:
            # 0-6 都是空的
            for str_name in str1[7:27]:
                if (str_name == "INIT:Mount_START") & (str_name in line) & ("INIT:Mount_START --late" not in line):
                    times = pattern3.findall(line)[1]
                    keyworldStr.append(str_name)
                    time.append(times)
                    break

                if line.find("INIT:Mount_START --late") != -1:
                    times = pattern3.findall(line)[1]
                    keyworldStr.append("INIT:Mount_START --late")
                    time.append(times)
                    break

                if (str_name == "INIT:Mount_END") & (str_name in line) & ("INIT:Mount_END --late" not in line):
                    times = pattern3.findall(line)[1]
                    keyworldStr.append(str_name)
                    time.append(times)
                    break

                if line.find("INIT:Mount_END --late") != -1:
                    times = pattern3.findall(line)[1]
                    keyworldStr.append("INIT:Mount_END --late")
                    time.append(times)
                    break

                if (str_name == "INIT:post-fs") & (str_name in line) & ("INIT:post-fs-data" not in line):
                    times = pattern3.findall(line)[1]
                    keyworldStr.append(str_name)
                    time.append(times)
                    break;

                if line.find("INIT:post-fs-data") != -1:
                    times = pattern3.findall(line)[1]
                    keyworldStr.append("INIT:post-fs-data")
                    time.append(times)
                    break;

                if str_name in line:
                    times = pattern3.findall(line)[1]
                    keyworldStr.append(str_name)
                    time.append(times)
                    break;

            for str_name in str1[27:39]:
                if str_name in line:
                    times = pattern4.findall(line)[-1]
                    keyworldStr.append(str_name)
                    time.append(times)
                    break;

    df = pd.DataFrame()
    df['Keyword'] = keyworldStr
    df['Time' + name] = time
    df = df.apply(np.roll, shift=3)
    # //制成了表格
    return df

def merge_op(df_contrast, df_test):
    #connect_df=pd.concat([df_contrast,df_test],axis=0);
    connect_df = pd.merge(df_contrast, df_test, how='inner', on='Keyword')
    connect_df['Timecont'] = connect_df['Timecont'].astype(np.float64)
    connect_df['Timebugreport'] = connect_df['Timebugreport'].astype(np.float64)
    for i in range(20):
       connect_df.iloc[i, 1] = float(connect_df.iloc[i, 1]) * 1000
       connect_df.iloc[i, 2] = float(connect_df.iloc[i, 2]) * 1000

    #减去之后的表格数据
    connect_new = connect_df;
    connect_new['contrast'] = connect_new['Timecont'].diff()
    connect_new['test'] = connect_new['Timebugreport'].diff()
    connect_new['contrast-test'] = connect_new['contrast'] - connect_new['test'];

    return connect_new;
def chuankou_create_workshhet(df):
    df.insert(loc=0, column='Boot Step', value=None)
    # df['Time'] = df['Time'].astype(str)
    # df['Time'] = df['Time'].apply(lambda x: x.strip("[]"))
    # df['Time'] = df['Time'].apply(lambda x: x.strip("''"))

    df.to_excel('chuankou_output.xlsx', index=False)

    wb = openpyxl.load_workbook('chuankou_output.xlsx')
    ws = wb.active
    # 创建并合并pl_lk单元格
    # ws.insert_cols(1,amount = 1) #row 行
    ws.cell(row=2, column=1, value='Pl_lk')
    ws.merge_cells(start_row=2, start_column=1, end_row=8, end_column=1)

    # 创建并合并kernel单元格
    ws.cell(row=9, column=1, value='Kernel')
    ws.merge_cells(start_row=9, start_column=1, end_row=28, end_column=1)

    # # 创建并合并Android单元格
    cell = ws.cell(row=29, column=1, value='Android')
    ws.merge_cells(start_row=29, start_column=1, end_row=46, end_column=1)
    wb.save("chuankou_output.xlsx")
# def mgerge_chuankou(table):
#     //connect_new['contrast'] = connect_new['Timecont'].diff()

if __name__ == "__main__":

    import argparse
    parser = argparse.ArgumentParser()
    parser.description = 'please enter two parameters a and b ...'
    parser.add_argument("-chuankou", "--chuankou", help="chuankou patch", type=str, default="")
    parser.add_argument("-bugreport", "--bugreport", help="bugreport patch", type=str, default="")
    parser.add_argument("-bugcont", "--bugcont", help="contrast patch", type=str, default="")
    parser.add_argument("-excel", "--excel", help="contrast patch", type=str, default="对比文件.xlsx")
    args = parser.parse_args()

    logger.info("keyworld find...")
    if args.chuankou != "":
       chuankoudf  = readkey(load_log_file('data/串口.log'),"chuankou",args.excel)
       flag = chuankoudf.Keyword.duplicated()
       if any(flag):
           # 有重复关键字
           print("串口log ,有重复关键字，请检测")
       else:
           chuankouf=chuankoudf['Timechuankou'];
           result1=chuankouf.str.split(':',expand=True);
           result2= result1[2].str.split('.',expand=True).astype(float)
           result1=result1.astype(float);
           result3=(result1[1]*60+ result2[0])*1000+ result2[1];
           chuankoudf['Time']=result3;
           chuankoudf["ChaTime"] = chuankoudf['Time'].diff()
           chuankou_create_workshhet(chuankoudf)
    if args.bugreport != "":
        bugreportdf = readkey(load_log_file(args.bugreport), "bugreport", args.excel);
        flagbug = bugreportdf.Keyword.duplicated()
        if any(flagbug):
            # 有重复关键字
            print("buglog ,有重复关键字，请检测")
    if args.bugcont != "":
       bugcontdf   = readkey(load_log_file(args.bugcont), "bugcont", args.excel);
       bugcontdf.columns = ["Keyword", 'Timecont']
       flagcon = bugcontdf.Keyword.duplicated()
       if any(flagcon):
           # 有重复关键字
           print("对比buglog ,有重复关键字，请检测")
    if  args.bugreport != "" and args.bugcont != "":
        df = merge_op(bugcontdf, bugreportdf);
        create_workshhet(df)




    # # 表格A 和 表格B
    # # df_contrast = find_keyword(results_contrast, "contrast");
    # # df_test = find_keyword(results_test, "test");
    #
    # # 合并表格并且操作;
    # df = merge_op(bugcontdf, bugreportdf);
    #
    # logger.info("create output xlsx start....")
    # create_workshhet(df);